"""
DENSO Visual Inspector — AI-Powered Anomaly Detection System
Colores corporativos DENSO: rojo #E30613, blanco, gris oscuro industrial
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import json
from pathlib import Path

# ── Librerías opcionales ──────────────────────────────────────────────────────
try:
    import numpy as np
    NUMPY_OK = True
except ImportError:
    NUMPY_OK = False

try:
    from PIL import Image, ImageTk
    PIL_OK = True
except ImportError:
    PIL_OK = False

try:
    import openvino as ov
    OPENVINO_OK = True
except ImportError:
    OPENVINO_OK = False

try:
    import cv2
    CV2_OK = True
except ImportError:
    CV2_OK = False

# ── Paleta corporativa DENSO ──────────────────────────────────────────────────
C = {
    "bg_dark":       "#1A1A1A",
    "bg_panel":      "#222222",
    "bg_card":       "#2A2A2A",
    "bg_hover":      "#333333",
    "bg_input":      "#1A1A1A",
    "denso_red":     "#E30613",
    "denso_red_dk":  "#B8000F",
    "ok_green":      "#2ECC71",
    "warning":       "#F39C12",
    "white":         "#FFFFFF",
    "text_primary":  "#F0F0F0",
    "text_secondary":"#A0A0A0",
    "text_dim":      "#606060",
    "border":        "#3A3A3A",
}

FONT_MONO  = ("Courier New", 9)
FONT_SMALL = ("Arial", 8)
IMG_EXTS   = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff"}

# ═════════════════════════════════════════════════════════════════════════════
# INFERENCIA
# ═════════════════════════════════════════════════════════════════════════════
def simulate_inference(image_path):
    import random, time
    time.sleep(0.06)
    score = random.uniform(0.0, 1.0)
    threshold = 0.5
    label = "NG" if score >= threshold else "OK"
    if PIL_OK and NUMPY_OK:
        w, h = 224, 224
        heat = np.zeros((h, w, 3), dtype=np.uint8)
        cx, cy = random.randint(50, 174), random.randint(50, 174)
        for y in range(h):
            for x in range(w):
                d = np.sqrt((x - cx)**2 + (y - cy)**2)
                v = max(0, 1 - d / 80)
                heat[y, x] = [int(v * 227), int(v * 6), int(v * 19)]
        heatmap_img = Image.fromarray(heat)
    else:
        heatmap_img = None
    return {
        "score": score,
        "threshold": threshold,
        "label": label,
        "heatmap": heatmap_img,
        "heatmap_gray": None,
        "mask": None,
        "combined": heatmap_img,
    }

def run_openvino_inference(model_path, image_path):
    #─────────────────────────────────────────────────────────────
    # Load model
    #─────────────────────────────────────────────────────────────
    core = ov.Core()
    model = core.read_model(str(model_path))
    compiled = core.compile_model(model, "CPU")

    #─────────────────────────────────────────────────────────────
    # Auto-detect outputs (model-agnostic)
    #─────────────────────────────────────────────────────────────
    outputs = list(compiled.outputs)
    output_names = [o.get_any_name() for o in outputs]
    print(f"[OpenVINO] Model outputs detected: {output_names}")

    # Try to find score output (could be named: pred_score, scores, output, etc.)
    score_output = None
    for name in output_names:
        lower = name.lower()
        if any(k in lower for k in ["score", "pred", "anomaly_score", "output"]):
            score_output = compiled.output(name)
            break
    if score_output is None:
        score_output = compiled.output(0)  # fallback to first output

    # Try to find anomaly map (could be: anomaly_map, map, pred_mask, etc.)
    map_output = None
    for name in output_names:
        lower = name.lower()
        if any(k in lower for k in ["map", "mask", "heatmap", "localization"]):
            map_output = compiled.output(name)
            break
    if map_output is None and len(outputs) > 1:
        map_output = compiled.output(1)  # fallback to second output

    # Try to find mask output (optional)
    mask_output = None
    for name in output_names:
        lower = name.lower()
        if "mask" in lower:
            mask_output = compiled.output(name)
            break

    #─────────────────────────────────────────────────────────────
    # Metadata threshold
    #─────────────────────────────────────────────────────────────
    model_dir = Path(model_path).parent
    meta_path = model_dir / "metadata.json"
    threshold = 0.5
    if meta_path.exists():
        with open(meta_path, "r") as f:
            meta = json.load(f)
        # Try multiple possible threshold keys
        threshold = meta.get("image_threshold",
            meta.get("pixel_threshold",
                meta.get("threshold",
                    meta.get("optimal_threshold", 0.5))))
        print(f"[OpenVINO] Threshold from metadata: {threshold}")

    #─────────────────────────────────────────────────────────────
    # Read image
    #─────────────────────────────────────────────────────────────
    img_bgr = cv2.imread(str(image_path))
    if img_bgr is None:
        raise RuntimeError(f"Could not read image: {image_path}")
    orig_h, orig_w = img_bgr.shape[:2]

    #─────────────────────────────────────────────────────────────
    # Auto-detect input size from model
    #─────────────────────────────────────────────────────────────
    input_layer = compiled.input(0)
    input_shape = input_layer.shape
    # Handle both static and dynamic shapes [?,3,H,W] or [1,3,H,W]
    if len(input_shape) == 4:
        H = int(input_shape[2]) if input_shape[2] > 0 else 256
        W = int(input_shape[3]) if input_shape[3] > 0 else 256
    else:
        H, W = 256, 256
    print(f"[OpenVINO] Input size: {H}x{W}")

    #─────────────────────────────────────────────────────────────
    # Preprocess (model-agnostic: resize + normalize + transpose)
    #─────────────────────────────────────────────────────────────
    img_resized = cv2.resize(img_bgr, (W, H))
    img_rgb = cv2.cvtColor(img_resized, cv2.COLOR_BGR2RGB)
    img_norm = img_rgb.astype(np.float32) / 255.0
    img_chw = np.transpose(img_norm, (2, 0, 1))
    img_batch = np.expand_dims(img_chw, axis=0)

    # Inference
    results = compiled({input_layer: img_batch})

    #─────────────────────────────────────────────────────────────
    # Extract score (handle various shapes)
    #─────────────────────────────────────────────────────────────
    score_raw = results[score_output]
    score = float(np.squeeze(score_raw))
    print(f"[OpenVINO] Raw score: {score}")

    #─────────────────────────────────────────────────────────────
    # Extract anomaly map (handle various shapes)
    #─────────────────────────────────────────────────────────────
    if map_output is not None:
        anomaly_map = results[map_output].squeeze()
        # Ensure 2D
        if len(anomaly_map.shape) > 2:
            anomaly_map = anomaly_map[0] if anomaly_map.shape[0] == 1 else anomaly_map
        # Ensure we have a 2D map
        if len(anomaly_map.shape) == 3 and anomaly_map.shape[2] in [1, 3]:
            anomaly_map = anomaly_map[:, :, 0]  # take first channel if multi-channel
    else:
        # Create dummy map if model has no map output
        anomaly_map = np.zeros((H, W), dtype=np.float32)

    #─────────────────────────────────────────────────────────────
    # Extract mask (optional)
    #─────────────────────────────────────────────────────────────
    if mask_output is not None:
        pred_mask = results[mask_output].squeeze().astype(np.uint8)
    else:
        pred_mask = np.zeros((H, W), dtype=np.uint8)

    #─────────────────────────────────────────────────────────────
    # Normalize anomaly map to 0-255
    #─────────────────────────────────────────────────────────────
    map_min = anomaly_map.min()
    map_max = anomaly_map.max()
    if map_max > map_min:
        heatmap = (
            (anomaly_map - map_min)
            / (map_max - map_min)
            * 255
        ).astype(np.uint8)
    else:
        heatmap = np.zeros((H, W), dtype=np.uint8)

    #─────────────────────────────────────────────────────────────
    # Resize outputs to original image size
    #─────────────────────────────────────────────────────────────
    heatmap_resized = cv2.resize(heatmap, (orig_w, orig_h))
    mask_resized = cv2.resize(
        pred_mask * 255,
        (orig_w, orig_h),
        interpolation=cv2.INTER_NEAREST
    )

    #─────────────────────────────────────────────────────────────
    # Visualization images
    #─────────────────────────────────────────────────────────────
    heatmap_bgr = cv2.cvtColor(heatmap_resized, cv2.COLOR_GRAY2BGR)
    heatmap_color = cv2.applyColorMap(heatmap_resized, cv2.COLORMAP_JET)
    mask_bgr = cv2.cvtColor(mask_resized, cv2.COLOR_GRAY2BGR)
    combined = np.hstack([img_bgr, heatmap_bgr, heatmap_color, mask_bgr])

    #─────────────────────────────────────────────────────────────
    # Label
    #─────────────────────────────────────────────────────────────
    label = "NG" if score >= threshold else "OK"
    color = (0, 0, 255) if label == "NG" else (0, 200, 0)
    cv2.putText(
        combined,
        f"{label} score={score:.4f}",
        (10, 30),
        cv2.FONT_HERSHEY_SIMPLEX,
        1.0,
        color,
        2
    )

    #─────────────────────────────────────────────────────────────
    # Convert for tkinter
    #─────────────────────────────────────────────────────────────
    heatmap_rgb = cv2.cvtColor(heatmap_color, cv2.COLOR_BGR2RGB)
    combined_rgb = cv2.cvtColor(combined, cv2.COLOR_BGR2RGB)
    mask_rgb = cv2.cvtColor(mask_bgr, cv2.COLOR_BGR2RGB)
    heatmap_gray_rgb = cv2.cvtColor(heatmap_bgr, cv2.COLOR_BGR2RGB)

    return {
        "score": score,
        "threshold": threshold,
        "label": label,
        "heatmap_color": Image.fromarray(heatmap_rgb),
        "heatmap_gray": Image.fromarray(heatmap_gray_rgb),
        "mask": Image.fromarray(mask_rgb),
        "combined": Image.fromarray(combined_rgb),
    }


def red_btn(parent, text, cmd, small=False):
    f = ("Arial", 9, "bold") if small else ("Arial", 10, "bold")
    b = tk.Button(parent, text=text, command=cmd,
        font=f, fg=C["white"], bg=C["denso_red"],
        activeforeground=C["white"], activebackground=C["denso_red_dk"],
        relief="flat", cursor="hand2",
        padx=10 if small else 16, pady=4 if small else 8)
    b.bind("<Enter>", lambda e: b.config(bg=C["denso_red_dk"]))
    b.bind("<Leave>", lambda e: b.config(bg=C["denso_red"]))
    return b

def ghost_btn(parent, text, cmd, small=False):
    f = ("Arial", 9) if small else ("Arial", 10)
    b = tk.Button(parent, text=text, command=cmd,
        font=f, fg=C["text_secondary"], bg=C["bg_hover"],
        activeforeground=C["white"], activebackground=C["bg_card"],
        relief="flat", cursor="hand2",
        padx=10 if small else 14, pady=4 if small else 7)
    b.bind("<Enter>", lambda e: b.config(fg=C["white"], bg=C["bg_card"]))
    b.bind("<Leave>", lambda e: b.config(fg=C["text_secondary"], bg=C["bg_hover"]))
    return b

def section_lbl(parent, num, text):
    f = tk.Frame(parent, bg=C["bg_card"])
    f.pack(fill="x", pady=(10, 2))
    tk.Label(f, text=num, font=("Arial", 8, "bold"),
        fg=C["denso_red"], bg=C["bg_card"]).pack(side="left", padx=(0, 6))
    tk.Label(f, text=text, font=("Arial", 9, "bold"),
        fg=C["text_secondary"], bg=C["bg_card"]).pack(side="left")

def mk_entry(parent, var, width=48):
    return tk.Entry(parent, textvariable=var, font=FONT_MONO, width=width,
        bg=C["bg_input"], fg=C["text_primary"],
        insertbackground=C["denso_red"], relief="flat",
        highlightbackground=C["border"], highlightcolor=C["denso_red"],
        highlightthickness=1)

# ═════════════════════════════════════════════════════════════════════════════
# PANTALLA DE CONFIGURACIÓN
# ═════════════════════════════════════════════════════════════════════════════
class SetupScreen(tk.Frame):
    def __init__(self, parent, on_start):
        super().__init__(parent, bg=C["bg_dark"])
        self.on_start      = on_start
        self.model_path    = tk.StringVar()
        self.input_folder  = tk.StringVar()
        self.output_folder = tk.StringVar()
        self.threshold     = tk.DoubleVar(value=0.0)
        self.sim_mode      = tk.BooleanVar(value=not (OPENVINO_OK and CV2_OK))
        self._build()

    def _build(self):
        tk.Frame(self, bg=C["denso_red"], height=6).pack(fill="x")

        titlebar = tk.Frame(self, bg=C["bg_panel"], height=64)
        titlebar.pack(fill="x")
        titlebar.pack_propagate(False)
        tk.Label(titlebar, text="DENSO BARCELONA", font=("Arial", 23, "bold"),
            fg=C["denso_red"], bg=C["bg_panel"]).pack(side="left", padx=(24, 4), pady=12)
        tk.Label(titlebar, text=" AnomalyEye",
            font=("Arial", 18), fg=C["white"], bg=C["bg_panel"]).pack(side="left", pady=12)
        tk.Label(titlebar, text="Crafting the Core",
            font=("Arial", 9, "italic"), fg=C["text_dim"], bg=C["bg_panel"]
        ).pack(side="right", padx=24)

        outer = tk.Frame(self, bg=C["bg_dark"])
        outer.pack(expand=True, fill="both")
        center = tk.Frame(outer, bg=C["bg_dark"])
        center.place(relx=0.5, rely=0.5, anchor="center")

        panel = Card(center, padx=36, pady=28)
        panel.pack(ipadx=10)

        # Sim mode checkbox
        sim_row = tk.Frame(panel, bg=C["bg_card"])
        sim_row.pack(fill="x", pady=(0, 10))
        tk.Checkbutton(sim_row,
            text="  ⚡  SIMULATION Mode  (no OpenVINO model required)",
            variable=self.sim_mode,
            font=("Arial", 10, "bold"),
            fg=C["warning"], bg=C["bg_card"],
            selectcolor=C["bg_dark"],
            activebackground=C["bg_card"],
            activeforeground=C["warning"],
            cursor="hand2",
            command=self._toggle_sim
        ).pack(side="left")

        tk.Frame(panel, bg=C["border"], height=1).pack(fill="x", pady=(0, 4))

        # 01 Modelo
        section_lbl(panel, "01", "DASH MODEL (OpenVINO format .xml)")
        row1 = tk.Frame(panel, bg=C["bg_card"])
        row1.pack(fill="x", pady=(4, 14))
        self.model_entry = mk_entry(row1, self.model_path, width=48)
        self.model_entry.pack(side="left", padx=(0, 8))
        self.model_btn = ghost_btn(row1, "Browse…", self._browse_model, small=True)
        self.model_btn.pack(side="left")

        # 02 Origen
        section_lbl(panel, "02", "INPUT IMAGE FOLDER  (source)")
        row2 = tk.Frame(panel, bg=C["bg_card"])
        row2.pack(fill="x", pady=(4, 14))
        mk_entry(row2, self.input_folder, width=48).pack(side="left", padx=(0, 8))
        ghost_btn(row2, "Browse…", self._browse_input, small=True).pack(side="left")

        # 03 Salida
        section_lbl(panel, "03", "OUTPUT FOLDER  (Excel + heatmaps)")
        row3 = tk.Frame(panel, bg=C["bg_card"])
        row3.pack(fill="x", pady=(4, 14))
        mk_entry(row3, self.output_folder, width=48).pack(side="left", padx=(0, 8))
        ghost_btn(row3, "Browse…", self._browse_output, small=True).pack(side="left")

        # 04 Threshold
        section_lbl(panel, "04", "THRESHOLD  OK / NG")
        thr_row = tk.Frame(panel, bg=C["bg_card"])
        thr_row.pack(fill="x", pady=(4, 18))
        tk.Label(thr_row, text="Threshold:",
            font=("Arial", 10, "bold"), fg=C["text_secondary"], bg=C["bg_card"]
        ).pack(side="left", padx=(0, 8))
        self.thr_entry = tk.Entry(thr_row, textvariable=self.threshold,
            font=("Arial", 14, "bold"), width=12,
            bg=C["bg_input"], fg=C["denso_red"],
            insertbackground=C["denso_red"], relief="flat",
            highlightbackground=C["border"], highlightcolor=C["denso_red"],
            highlightthickness=1, justify="center")
        self.thr_entry.pack(side="left", padx=(0, 12))
        tk.Label(thr_row, text="Score < threshold → OK   |   Score ≥ threshold → NG",
            font=("Arial", 8), fg=C["text_dim"], bg=C["bg_card"]
        ).pack(side="left", padx=(6, 0))

        tk.Frame(panel, bg=C["border"], height=1).pack(fill="x", pady=(4, 16))
        red_btn(panel, "▶   START INSPECTION", self._start).pack()

        # Dependencias
        dep_row = tk.Frame(center, bg=C["bg_dark"])
        dep_row.pack(pady=(14, 0))
        for name, ok in [("numpy", NUMPY_OK), ("Pillow", PIL_OK),
                          ("OpenVINO", OPENVINO_OK), ("OpenCV", CV2_OK)]:
            tk.Label(dep_row,
                text=f"{'●' if ok else '○'} {name}",
                font=("Arial", 8),
                fg=C["ok_green"] if ok else C["text_dim"],
                bg=C["bg_dark"]).pack(side="left", padx=10)

        self._toggle_sim()

    def _toggle_sim(self):
        st = "disabled" if self.sim_mode.get() else "normal"
        self.model_entry.config(state=st)
        self.model_btn.config(state=st)

    def _browse_model(self):
        p = filedialog.askopenfilename(title="OpenVINO Model",
            filetypes=[("OpenVINO XML", "*.xml"), ("Todos", "*.*")])
        if p: self.model_path.set(p)

    def _browse_input(self):
        p = filedialog.askdirectory(title="Image folder")
        if p: self.input_folder.set(p)

    def _browse_output(self):
        p = filedialog.askdirectory(title="Output folder")
        if p: self.output_folder.set(p)

    def _start(self):
        if not self.input_folder.get():
            messagebox.showerror("Error", "Please select the input image folder.")
            return
        if not self.output_folder.get():
            messagebox.showerror("Error", "Please select the output folder.")
            return
        if not self.sim_mode.get() and OPENVINO_OK and CV2_OK and not self.model_path.get():
            messagebox.showerror("Error", "Please select an OpenVINO model (.xml).")
            return
        self.on_start(
            model_path    = self.model_path.get(),
            input_folder  = self.input_folder.get(),
            output_folder = self.output_folder.get(),
            threshold     = self.threshold.get(),
            sim_mode      = self.sim_mode.get(),
        )

# ═════════════════════════════════════════════════════════════════════════════
# PANTALLA DE RESULTADOS
# ═════════════════════════════════════════════════════════════════════════════
class ResultsScreen(tk.Frame):
    def __init__(self, parent, config, on_back):
        super().__init__(parent, bg=C["bg_dark"])
        self.config_data        = config
        self.on_back            = on_back
        self.threshold          = tk.DoubleVar(value=config["threshold"])
        self.results            = []
        self.filtered           = []   # lista de índices en self.results visibles
        self.current_result_idx = None
        self.filter_mode        = tk.StringVar(value="TODOS")
        self.search_var         = tk.StringVar()
        self._filter_btns       = {}
        self._build()
        self.after(100, self._start_inference)

    # ─────────────────────────────────────────────────────────────────────────
    def _build(self):
        tk.Frame(self, bg=C["denso_red"], height=4).pack(fill="x")

        topbar = tk.Frame(self, bg=C["bg_panel"], height=50)
        topbar.pack(fill="x")
        topbar.pack_propagate(False)
        ghost_btn(topbar, "◀  Back", self.on_back, small=True
            ).pack(side="left", padx=10, pady=10)
        tk.Label(topbar, text="DENSO BARCELONA", font=("Arial", 13, "bold"),
            fg=C["denso_red"], bg=C["bg_panel"]).pack(side="left", padx=(4, 2))
        tk.Label(topbar, text="Inference Tab — Results",
            font=("Arial", 11), fg=C["white"], bg=C["bg_panel"]).pack(side="left")
        red_btn(topbar, "⬇  Export Excel", self._export_excel, small=True
            ).pack(side="right", padx=10, pady=10)

        body = tk.Frame(self, bg=C["bg_dark"])
        body.pack(fill="both", expand=True, padx=12, pady=10)

        left = tk.Frame(body, bg=C["bg_dark"], width=295)
        left.pack(side="left", fill="y", padx=(0, 10))
        left.pack_propagate(False)
        self._build_left(left)

        right = tk.Frame(body, bg=C["bg_dark"])
        right.pack(side="left", fill="both", expand=True)
        self._build_right(right)

    # ── Panel izquierdo ────────────────────────────────────────────────────
    def _build_left(self, parent):
        # Stats
        stats = Card(parent, padx=12, pady=10)
        stats.pack(fill="x", pady=(0, 8))
        tk.Label(stats, text="SUMMARY", font=("Arial", 7, "bold"),
            fg=C["text_dim"], bg=C["bg_card"]).pack(anchor="w")
        row = tk.Frame(stats, bg=C["bg_card"])
        row.pack(fill="x", pady=6)
        self.lbl_ok    = self._stat(row, "0", "OK",    C["ok_green"])
        self.lbl_ng    = self._stat(row, "0", "NG",    C["denso_red"])
        self.lbl_total = self._stat(row, "0", "TOTAL", C["text_secondary"])
        for w in (self.lbl_ok, self.lbl_ng, self.lbl_total):
            w.pack(side="left", expand=True)
        self.progress_lbl = tk.Label(stats, text="Starting…",
            font=("Arial", 8), fg=C["text_secondary"], bg=C["bg_card"])
        self.progress_lbl.pack(anchor="w")
        self.progress_var = tk.DoubleVar()
        pb = ttk.Progressbar(stats, variable=self.progress_var, maximum=100, length=260)
        pb.pack(fill="x", pady=(3, 0))
        s = ttk.Style(); s.theme_use("default")
        s.configure("TProgressbar", troughcolor=C["bg_dark"],
            background=C["denso_red"], thickness=5)

        # Threshold live
        thr_card = Card(parent, padx=12, pady=10)
        thr_card.pack(fill="x", pady=(0, 8))
        tk.Label(thr_card, text="THRESHOLD  (live adjustment)",
            font=("Arial", 7, "bold"), fg=C["text_dim"], bg=C["bg_card"]
        ).pack(anchor="w", pady=(0, 4))
        thr_input_row = tk.Frame(thr_card, bg=C["bg_card"])
        thr_input_row.pack(fill="x", pady=(0, 4))
        self.thr_display = tk.Entry(thr_input_row, textvariable=self.threshold,
            font=("Arial", 16, "bold"), width=14,
            bg=C["bg_input"], fg=C["denso_red"],
            insertbackground=C["denso_red"], relief="flat",
            highlightbackground=C["border"], highlightcolor=C["denso_red"],
            highlightthickness=1, justify="center")
        self.thr_display.pack(side="left")
        tk.Label(thr_input_row, text="  ← Press Enter to apply",
            font=("Arial", 8), fg=C["text_dim"], bg=C["bg_card"]
        ).pack(side="left", padx=(6, 0))
        self.thr_display.bind("<Return>", lambda e: self._on_threshold_change())
        self.thr_display.bind("<FocusOut>", lambda e: self._on_threshold_change())

        # ── Filtro OK / NG / TODOS ─────────────────────────────────────────
        filter_card = Card(parent, padx=12, pady=10)
        filter_card.pack(fill="x", pady=(0, 8))
        tk.Label(filter_card, text="FILTER LIST",
            font=("Arial", 7, "bold"), fg=C["text_dim"], bg=C["bg_card"]
        ).pack(anchor="w", pady=(0, 6))
        btn_row = tk.Frame(filter_card, bg=C["bg_card"])
        btn_row.pack(fill="x")
        for label in ("TODOS", "OK", "NG"):
            b = tk.Button(btn_row, text=label,
                font=("Arial", 9, "bold"),
                fg=C["text_primary"], bg=C["bg_hover"],
                relief="flat", cursor="hand2",
                padx=14, pady=5,
                command=lambda l=label: self._apply_filter(l))
            b.pack(side="left", padx=(0, 4))
            self._filter_btns[label] = b
        self._highlight_filter_btn("TODOS")

        # ── Buscador ───────────────────────────────────────────────────────
        search_card = Card(parent, padx=12, pady=10)
        search_card.pack(fill="x", pady=(0, 8))
        tk.Label(search_card, text="SEARCH IMAGE",
            font=("Arial", 7, "bold"), fg=C["text_dim"], bg=C["bg_card"]
        ).pack(anchor="w", pady=(0, 4))
        srow = tk.Frame(search_card, bg=C["bg_card"])
        srow.pack(fill="x")
        self.search_entry = mk_entry(srow, self.search_var, width=20)
        self.search_entry.pack(side="left", fill="x", expand=True, padx=(0, 6))
        ghost_btn(srow, "✕", self._clear_search, small=True).pack(side="left")
        self.search_var.trace_add("write", lambda *_: self._apply_filter(self.filter_mode.get()))

        # Tip
        tk.Label(search_card, text="Right-click on list → copy filename",
            font=("Arial", 7), fg=C["text_dim"], bg=C["bg_card"]
        ).pack(anchor="w", pady=(4, 0))

        # ── Lista ──────────────────────────────────────────────────────────
        list_card = Card(parent, padx=8, pady=8)
        list_card.pack(fill="both", expand=True)
        hdr = tk.Frame(list_card, bg=C["bg_card"])
        hdr.pack(fill="x", pady=(0, 4))
        tk.Label(hdr, text="IMAGES", font=("Arial", 7, "bold"),
            fg=C["text_dim"], bg=C["bg_card"]).pack(side="left", padx=4)
        self.list_count_lbl = tk.Label(hdr, text="",
            font=("Arial", 7), fg=C["text_dim"], bg=C["bg_card"])
        self.list_count_lbl.pack(side="right", padx=4)

        lbf = tk.Frame(list_card, bg=C["bg_card"])
        lbf.pack(fill="both", expand=True)
        sb = tk.Scrollbar(lbf, bg=C["bg_dark"], troughcolor=C["bg_panel"])
        sb.pack(side="right", fill="y")
        self.listbox = tk.Listbox(lbf, font=FONT_MONO, selectmode="single",
            bg=C["bg_input"], fg=C["text_primary"],
            selectbackground=C["denso_red"], selectforeground=C["white"],
            relief="flat", highlightthickness=0,
            yscrollcommand=sb.set, activestyle="none")
        self.listbox.pack(fill="both", expand=True)
        sb.config(command=self.listbox.yview)
        self.listbox.bind("<<ListboxSelect>>", self._on_list_select)

        # Menú contextual
        self.ctx_menu = tk.Menu(self, tearoff=0,
            bg=C["bg_card"], fg=C["text_primary"],
            activebackground=C["denso_red"], activeforeground=C["white"],
            relief="flat")
        self.ctx_menu.add_command(label="📋  Copy filename",
            command=self._copy_name)
        self.ctx_menu.add_command(label="📋  Copy full path",
            command=self._copy_path)
        self.listbox.bind("<Button-3>", self._show_ctx_menu)

    # ── Panel derecho ──────────────────────────────────────────────────────
    def _build_right(self, parent):
        vh = tk.Frame(parent, bg=C["bg_panel"], height=44)
        vh.pack(fill="x")
        vh.pack_propagate(False)

        self.viewer_filename = tk.Label(vh, text="—",
            font=("Arial", 11, "bold"), fg=C["white"], bg=C["bg_panel"])
        self.viewer_filename.pack(side="left", padx=12, pady=8)

        ghost_btn(vh, "📋 Copy name", self._copy_name, small=True
            ).pack(side="left", padx=(0, 8), pady=8)

        self.verdict_lbl = tk.Label(vh, text="",
            font=("Arial", 11, "bold"), bg=C["bg_panel"])
        self.verdict_lbl.pack(side="right", padx=8)
        self.score_lbl = tk.Label(vh, text="",
            font=("Arial", 10), fg=C["text_secondary"], bg=C["bg_panel"])
        self.score_lbl.pack(side="right", padx=4)

        img_zone = tk.Frame(parent, bg=C["bg_dark"])
        img_zone.pack(fill="both", expand=True, pady=8)

        orig_card = Card(img_zone, padx=4, pady=4)
        orig_card.pack(side="left", fill="both", expand=True, padx=(0, 6))
        tk.Label(orig_card, text="ORIGINAL", font=("Arial", 7, "bold"),
            fg=C["text_dim"], bg=C["bg_card"]).pack(anchor="w")
        self.orig_canvas = tk.Canvas(orig_card, bg=C["bg_dark"], highlightthickness=0)
        self.orig_canvas.pack(fill="both", expand=True)

        heat_card = Card(img_zone, padx=4, pady=4)
        heat_card.pack(side="left", fill="both", expand=True)
        tk.Label(heat_card, text="HEATMAP", font=("Arial", 7, "bold"),
            fg=C["text_dim"], bg=C["bg_card"]).pack(anchor="w")
        self.heat_canvas = tk.Canvas(heat_card, bg=C["bg_dark"], highlightthickness=0)
        self.heat_canvas.pack(fill="both", expand=True)

        nav = tk.Frame(parent, bg=C["bg_dark"])
        nav.pack(fill="x", pady=(0, 4))
        ghost_btn(nav, "◀ Previous", lambda: self._navigate(-1), small=True
            ).pack(side="left", padx=4)
        self.nav_lbl = tk.Label(nav, text="",
            font=("Arial", 9), fg=C["text_secondary"], bg=C["bg_dark"])
        self.nav_lbl.pack(side="left", expand=True)
        ghost_btn(nav, "Next ▶", lambda: self._navigate(1), small=True
            ).pack(side="right", padx=4)

    def _stat(self, parent, value, label, color):
        f = tk.Frame(parent, bg=C["bg_card"])
        tk.Label(f, text=value, font=("Arial", 18, "bold"),
            fg=color, bg=C["bg_card"]).pack()
        tk.Label(f, text=label, font=("Arial", 7, "bold"),
            fg=C["text_dim"], bg=C["bg_card"]).pack()
        return f

    # ─────────────────────────────────────────────────────────────────────────
    # INFERENCIA
    # ─────────────────────────────────────────────────────────────────────────
    def _start_inference(self):
        folder = Path(self.config_data["input_folder"])
        images = sorted([p for p in folder.iterdir() if p.suffix.lower() in IMG_EXTS])
        if not images:
            messagebox.showwarning("No images found", "No images found in the selected folder.")
            return
        threading.Thread(target=self._infer_all, args=(images,), daemon=True).start()

    def _infer_all(self, images):
        total = len(images)
        model = self.config_data["model_path"]
        out   = self.config_data["output_folder"]
        sim   = self.config_data.get("sim_mode", False)

        for i, img_path in enumerate(images):
            try:
                if not sim and OPENVINO_OK and CV2_OK and model:
                    result = run_openvino_inference(model, img_path)
                    score = result["score"]
                    heatmap = result["heatmap_color"]
                    label = result["label"]
                    combined = result["combined"]
                    real_threshold = result["threshold"]
                    heatmap_gray = result["heatmap_gray"]
                    mask = result["mask"]
                else:
                    result = simulate_inference(img_path)
                    score = result["score"]
                    heatmap = result["heatmap"]
                    label = result["label"]
                    combined = result["combined"]
                    real_threshold = result["threshold"]
                    heatmap_gray = None
                    mask = None
            except Exception as e:
                print(e)
                score = 0.0
                heatmap = None
                combined = None
                label = "ERROR"
                real_threshold = 0.5
                heatmap_gray = None
                mask = None

            hm_path = None
            if combined and out:
                hm_path = Path(out) / f"{label}_{img_path.stem}_heatmap.png"
                try:
                    combined.save(hm_path)
                except Exception:
                    pass

            self.results.append({
                "path": img_path,
                "score": score,
                "label": label,
                "threshold": real_threshold,
                "heatmap": heatmap,
                "heatmap_gray": heatmap_gray,
                "mask": mask,
                "combined": combined,
                "hm_path": hm_path,
            })
            self.after(0, self._update_progress, i + 1, total,
                       (i + 1) / total * 100, img_path.name)

        self.after(0, self._inference_done)

    def _update_progress(self, done, total, pct, name):
        self.progress_var.set(pct)
        self.progress_lbl.config(text=f"Processing {done}/{total}: {name}")
        self._refresh_stats()

    def _inference_done(self):
        self.progress_lbl.config(text="✔  Inference complete")
        self._refresh_stats()
        self._apply_filter("TODOS")

    # ─────────────────────────────────────────────────────────────────────────
    # FILTRO + BÚSQUEDA
    # ─────────────────────────────────────────────────────────────────────────
    def _apply_filter(self, mode):
        self.filter_mode.set(mode)
        self._highlight_filter_btn(mode)
        thr   = self.threshold.get()
        query = self.search_var.get().lower().strip()
        self.filtered = [
            idx for idx, r in enumerate(self.results)
            if (mode == "TODOS" or (mode == "OK") == (r["score"] < thr))
            and (not query or query in r["path"].name.lower())
        ]
        self._refresh_list()

    def _highlight_filter_btn(self, active):
        for label, btn in self._filter_btns.items():
            if label == active:
                btn.config(bg=C["denso_red"], fg=C["white"])
            elif label == "OK":
                btn.config(bg=C["bg_hover"], fg=C["ok_green"])
            elif label == "NG":
                btn.config(bg=C["bg_hover"], fg=C["denso_red"])
            else:
                btn.config(bg=C["bg_hover"], fg=C["text_primary"])

    def _clear_search(self):
        self.search_var.set("")
        self.search_entry.focus()

    # ─────────────────────────────────────────────────────────────────────────
    # LISTA
    # ─────────────────────────────────────────────────────────────────────────
    def _refresh_list(self):
        thr = self.threshold.get()
        self.listbox.delete(0, "end")
        for fi in self.filtered:
            r   = self.results[fi]
            ok  = r["score"] < thr
            self.listbox.insert("end",
                f" {'✓' if ok else '✗'}  {r['path'].name}  [{r['score']:.3f}]")
            self.listbox.itemconfig(self.listbox.size() - 1,
                fg=C["ok_green"] if ok else C["denso_red"])
        self.list_count_lbl.config(
            text=f"{len(self.filtered)} / {len(self.results)}")
        if self.filtered:
            self._show_result(self.filtered[0])

    def _refresh_stats(self):
        thr = self.threshold.get()
        ok  = sum(1 for r in self.results if r["score"] < thr)
        self.lbl_ok.winfo_children()[0].config(text=str(ok))
        self.lbl_ng.winfo_children()[0].config(text=str(len(self.results) - ok))
        self.lbl_total.winfo_children()[0].config(text=str(len(self.results)))

    def _on_threshold_change(self, event=None):
        if self.results:
            self._refresh_stats()
            self._apply_filter(self.filter_mode.get())

    def _on_list_select(self, event):
        sel = self.listbox.curselection()
        if sel:
            self._show_result(self.filtered[sel[0]])

    def _navigate(self, delta):
        if not self.filtered:
            return
        try:
            pos = self.filtered.index(self.current_result_idx)
        except (ValueError, TypeError):
            pos = 0
        pos = (pos + delta) % len(self.filtered)
        fi  = self.filtered[pos]
        self._show_result(fi)
        self.listbox.selection_clear(0, "end")
        self.listbox.selection_set(pos)
        self.listbox.see(pos)

    # ─────────────────────────────────────────────────────────────────────────
    # VISOR
    # ─────────────────────────────────────────────────────────────────────────
    def _show_result(self, result_idx):
        self.current_result_idx = result_idx
        r   = self.results[result_idx]
        thr = self.threshold.get()
        ok  = r["score"] < thr
        try:
            pos = self.filtered.index(result_idx)
        except ValueError:
            pos = 0

        self.viewer_filename.config(text=r["path"].name, fg=C["white"])
        self.score_lbl.config(text=f"Score: {r['score']:.4f}")
        self.verdict_lbl.config(
            text="●  OK" if ok else "●  NG",
            fg=C["ok_green"] if ok else C["denso_red"])
        self.nav_lbl.config(text=f"{pos + 1} / {len(self.filtered)}")

        self._draw_image(self.orig_canvas, r["path"])
        if r["heatmap"]:
            self._draw_pil(self.heat_canvas, r["heatmap"])
        else:
            self.heat_canvas.delete("all")
            self.heat_canvas.create_text(150, 150,
                text="No heatmap", fill=C["text_dim"], font=FONT_SMALL)

    def _draw_image(self, canvas, path):
        if not PIL_OK: return
        try: self._blit(canvas, Image.open(path))
        except Exception: canvas.delete("all")

    def _draw_pil(self, canvas, img):
        self._blit(canvas, img)

    def _blit(self, canvas, pil_img):
        canvas.update_idletasks()
        cw = canvas.winfo_width()  or 320
        ch = canvas.winfo_height() or 320
        img = pil_img.copy()
        img.thumbnail((cw, ch), Image.LANCZOS)
        tk_img = ImageTk.PhotoImage(img)
        canvas.delete("all")
        canvas.create_image(cw // 2, ch // 2, anchor="center", image=tk_img)
        canvas._tk_img = tk_img

    # ─────────────────────────────────────────────────────────────────────────
    # COPIAR NOMBRE / RUTA
    # ─────────────────────────────────────────────────────────────────────────
    def _current_path(self):
        if self.current_result_idx is not None:
            return self.results[self.current_result_idx]["path"]
        return None

    def _copy_name(self):
        p = self._current_path()
        if not p: return
        self.clipboard_clear(); self.clipboard_append(p.name)
        self._flash(f"✔  Copied: {p.name}")

    def _copy_path(self):
        p = self._current_path()
        if not p: return
        self.clipboard_clear(); self.clipboard_append(str(p))
        self._flash("✔  Full path copied")

    def _flash(self, msg):
        orig = self.viewer_filename.cget("text")
        self.viewer_filename.config(text=msg, fg=C["ok_green"])
        self.after(1800, lambda: self.viewer_filename.config(
            text=orig, fg=C["white"]))

    def _show_ctx_menu(self, event):
        sel = self.listbox.nearest(event.y)
        if sel >= 0:
            self.listbox.selection_clear(0, "end")
            self.listbox.selection_set(sel)
            self._show_result(self.filtered[sel])
        try:
            self.ctx_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.ctx_menu.grab_release()

    # ─────────────────────────────────────────────────────────────────────────
    # EXPORTAR
    # ─────────────────────────────────────────────────────────────────────────
    def _export_excel(self):
        if not self.results:
            messagebox.showinfo("No data", "No results yet.")
            return
        out = Path(self.config_data["output_folder"])
        thr = self.threshold.get()

        # CSV (siempre)
        csv_path = out / "inspection_results.csv"
        try:
            lines = ["Imagen,Score,Resultado,Heatmap\n"]
            for r in self.results:
                ok = r["score"] < thr
                lines.append(
                    f"{r['path'].name},{r['score']:.6f},{'OK' if ok else 'NG'},"
                    f"{r['hm_path'] or ''}\n")
            csv_path.write_text("".join(lines), encoding="utf-8")
        except Exception as e:
            messagebox.showerror("CSV Error", str(e)); return

        # XLSX con colores DENSO
        xlsx_ok = False
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"
            # Cabecera con fondo rojo DENSO
            headers = ["Image", "Score", "Result", "Heatmap", "Threshold"]
            ws.append(headers)
            for col in range(1, 6):
                cell = ws.cell(1, col)
                cell.fill = PatternFill("solid", fgColor="E30613")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            # Datos
            for r in self.results:
                ok = r["score"] < thr
                ws.append([r["path"].name, round(r["score"], 6),
                    "OK" if ok else "NG", str(r["hm_path"] or ""), thr])
                cell = ws.cell(ws.max_row, 3)
                cell.fill = PatternFill("solid", fgColor="2ECC71" if ok else "E30613")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 12
            wb.save(out / "inspection_results.xlsx")
            xlsx_ok = True
        except ImportError:
            pass

        msg = f"CSV guardado:\n{csv_path}"
        if xlsx_ok:
            msg += "\n\nExcel File (.xlsx) Generated on selected address"
        messagebox.showinfo("Exported", msg)

# ═════════════════════════════════════════════════════════════════════════════
# APP
# ═════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("DENSO BARCELONA AnomalyEye")
        self.geometry("1200x750")
        self.minsize(980, 640)
        self.configure(bg=C["bg_dark"])
        try:
            if PIL_OK:
                ico = Image.new("RGB", (32, 32), color=(227, 6, 19))
                self._ico = ImageTk.PhotoImage(ico)
                self.iconphoto(True, self._ico)
        except Exception:
            pass
        self._show_setup()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    def _show_setup(self):
        self._clear()
        SetupScreen(self, on_start=self._show_results).pack(fill="both", expand=True)

    def _show_results(self, **config):
        self._clear()
        ResultsScreen(self, config=config, on_back=self._show_setup
            ).pack(fill="both", expand=True)

if __name__ == "__main__":
    App().mainloop()
